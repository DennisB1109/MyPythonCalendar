from calendar import month
from cgitb import text
from http.server import executable
from os import system
from os import path


from datetime import *
from datetime import timedelta
from sre_parse import State

from uuid import uuid1                                          # To assign each event a unique ID

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from tkinter import messagebox
from tkinter import *
from tkcalendar import *

import time
from winotify import Notification, audio

print("LETS GET STARTED\n")

def get_curr_day():
    """_summary_

    Returns:
        _type_: _description_
    """
    convert_to_german = {"Monday" : "Montag", "Tuesday" : "Dienstag", "Wednesday" : "Mittwoch", "Thursday" : "Donnerstag", "Friday": "Freitag", "Saturday" : "Samstag", "Sunday" : "Sonntag"}
    weekday = datetime.datetime.now().strftime("%A")
    print("Heute ist der %ste, das ist ein" % (datetime.datetime.now().day), (convert_to_german.get(weekday)))
    return weekday

def get_curr_month():
    """_summary_

    Returns:
        _type_: _description_
    """
    month_name = {1 : "Januar", 2: "Februar", 3: "März", 4: "April", 5: "Mai", 6: "Juni", 7: "Juli", 8: "August", 9: "September", 10: "Oktober", 11: "November", 12: "Dezember"}
    number_of_month = datetime.datetime.now().month
    print("Wir haben", month_name.get(number_of_month), "(%s)" % (number_of_month))
    return number_of_month
    

def get_curr_year():
    """_summary_

    Returns:
        _type_: _description_
    """
    print("Wir haben das Jahr:", datetime.datetime.now().year)
    return datetime.datetime.now().year

def get_date():
    """_summary_

    Returns:
        _type_: _description_
    """
    print("Heutiges datum: %s-%s-%s" % (datetime.now().day, datetime.now().month, datetime.now().year))
    return f'{datetime.now().day}.{datetime.now().month}.{datetime.now().year}'

def add_event(day: str, month: str, year: str, description: str, reminder: str):
    """_summary_

    Returns:
        _type_: _description_
    """
    wb = load_workbook('my_Events.xlsx')
    ws = wb["all_Events"]
    uid = str(uuid1().int>>64)
    if len(year) == 2:
        temp = str(datetime.now().year)
        century = temp[0] + temp[1] + year
        temp = ''.join(century)
        year = temp
    ws.append([uid, day, month, year, description, reminder])
    wb.save('my_Events.xlsx')
    added_event = f"Your event for the {day}.{month}.{year} was successfully added to your Calendar"
    return added_event

def del_event(uid: str):
    """_summary_

    Returns:
        _type_: _description_
    """
    wb = load_workbook('my_Events.xlsx')
    ws = wb["all_Events"]
    temp_list = []
    was_event_deleted = False
    for row in range(2,100):
        temp_list.clear()
        temp_list.append(ws[get_column_letter(1) + str(row)].value)
        if temp_list[0] == uid:
            deleted_day = ws[get_column_letter(2) + str(row)].value
            deleted_month =ws[get_column_letter(3) + str(row)].value
            deleted_year =ws[get_column_letter(4) + str(row)].value
            print(f"Your Event for the {deleted_day}.{deleted_month}.{deleted_year} with the ID {uid} was successfully deleted")
            ws.delete_rows(row)
            wb.save('my_Events.xlsx')
            was_event_deleted = True
            break
    return uid

def show_all_events():
    """_summary_

    Returns:
        _type_: _description_
    """
    wb = load_workbook('my_Events.xlsx')
    ws = wb["all_Events"]    
    temp_list = []
    for row in range(3,100):
        check_empty = ws[get_column_letter(1) + str(row)].value
        if check_empty is None:
            break
        for col in range(1,7):
            temp_list.append(ws[get_column_letter(col) + str(row)].value)
    return temp_list

def check_todays_events():
    upcoming_event_description_list = []
    wb = load_workbook('my_Events.xlsx')
    ws = wb["all_Events"]
    for row in range(3, 100):
        get_event_day = ws[get_column_letter(2) + str(row)].value
        get_event_month = ws[get_column_letter(3) + str(row)].value
        get_event_year = ws[get_column_letter(4) + str(row)].value
        get_event_description = ws[get_column_letter(5) + str(row)].value
        date_event_string = f'{get_event_year}-{get_event_month}-{get_event_day}'

        temp_datetime_now = str(datetime.now())[:-16].strip()
        if temp_datetime_now == date_event_string:
            upcoming_event_description_list.append(get_event_description)
    return upcoming_event_description_list

def check_upcoming_events():
    """_summary_

    Returns:
        _type_: _description_
    """
    upcoming_event_date_list = []
    upcoming_event_description_list = []
    upcoming_event_reminder_list = []
    wb = load_workbook('my_Events.xlsx')
    ws = wb["all_Events"]
    for row in range(3, 100):
        save_reminder = ws[get_column_letter(6) + str(row)].value
        if save_reminder is None:
            break
        day = timedelta(days = int(save_reminder))
        get_event_day = ws[get_column_letter(2) + str(row)].value
        get_event_month = ws[get_column_letter(3) + str(row)].value
        get_event_year = ws[get_column_letter(4) + str(row)].value[2] + ws[get_column_letter(4) + str(row)].value[3]
        get_event_description = ws[get_column_letter(5) + str(row)].value
        date_event_string = f'{get_event_day}/{get_event_month}/{get_event_year} 00:00:01'
        date_time_obj = datetime.strptime(date_event_string, '%d/%m/%y %H:%M:%S')
        
        remind_on_this_date = str(date_time_obj - day)

        remind_on_this_date = remind_on_this_date[:-9].strip()
        temp_datetime_now = str(datetime.now())[:-16].strip()
        
        if temp_datetime_now == remind_on_this_date:
            date_event_string = date_event_string[:len(date_event_string) - 9]
            date_event_string = date_event_string.replace(r"/", ".")
            upcoming_event_date_list.append(date_event_string)
            upcoming_event_description_list.append(get_event_description)
            upcoming_event_reminder_list.append(save_reminder)
    return upcoming_event_description_list, upcoming_event_reminder_list

def write_past_events():
    wb = load_workbook('my_Events.xlsx')
    ws_all_events = wb["all_Events"]
    ws_past_events = wb["past_Events"]
    ws_past_events.delete_rows(3, 100)
    wb.save('my_Events.xlsx')
    
    for row in range(3, 100):
        if ws_all_events[get_column_letter(1) + str(row)].value is None:
            break
        get_event_id = ws_all_events[get_column_letter(1) + str(row)].value
        get_event_day = ws_all_events[get_column_letter(2) + str(row)].value
        get_event_month = ws_all_events[get_column_letter(3) + str(row)].value
        get_event_year = ws_all_events[get_column_letter(4) + str(row)].value[2] + ws_all_events[get_column_letter(4) + str(row)].value[3]
        get_event_description = ws_all_events[get_column_letter(5) + str(row)].value
        get_event_reminder = ws_all_events[get_column_letter(6) + str(row)].value
        
        date_event_string = f'{get_event_day}/{get_event_month}/{get_event_year} 00:00:01'
        date_time_obj = datetime.strptime(date_event_string, '%d/%m/%y %H:%M:%S')
        
        curr_date = datetime.now()

        if date_time_obj < curr_date:
            ws_past_events.append([get_event_id, get_event_day, get_event_month, get_event_year, get_event_description, get_event_reminder])
            wb.save('my_Events.xlsx')

def show_past_events() -> list:
    """_summary_

    Returns:
        _type_: _description_
    """
    wb = load_workbook('my_Events.xlsx')
    ws = wb["past_Events"]
    temp_list = []
    for row in range(3,100):
        check_empty = ws[get_column_letter(1) + str(row)].value
        if check_empty is None:
            break
        for col in range(1,7):
            temp_list.append(ws[get_column_letter(col) + str(row)].value)
    print(temp_list)
    return temp_list

def write_upcoming_events():
    wb = load_workbook('my_Events.xlsx')
    ws_all_events = wb["all_Events"]
    ws_upcoming_events = wb["upcoming_Events"]
    ws_upcoming_events.delete_rows(3, 100)
    wb.save('my_Events.xlsx')
    
    for row in range(3, 100):
        if ws_all_events[get_column_letter(1) + str(row)].value is None:
            break
        get_event_id = ws_all_events[get_column_letter(1) + str(row)].value
        get_event_day = ws_all_events[get_column_letter(2) + str(row)].value
        get_event_month = ws_all_events[get_column_letter(3) + str(row)].value
        get_event_year = ws_all_events[get_column_letter(4) + str(row)].value[2] + ws_all_events[get_column_letter(4) + str(row)].value[3]
        get_event_description = ws_all_events[get_column_letter(5) + str(row)].value
        get_event_reminder = ws_all_events[get_column_letter(6) + str(row)].value
        
        date_event_string = f'{get_event_day}/{get_event_month}/{get_event_year} 00:00:01'
        date_time_obj = datetime.strptime(date_event_string, '%d/%m/%y %H:%M:%S')
        
        curr_date = datetime.now()

        if date_time_obj >= curr_date:
            ws_upcoming_events.append([get_event_id, get_event_day, get_event_month, get_event_year, get_event_description, get_event_reminder])
            wb.save('my_Events.xlsx')

def show_upcoming_events() -> list:
    """_summary_

    Returns:
        _type_: _description_
    """
    wb = load_workbook('my_Events.xlsx')
    ws = wb["upcoming_Events"]
    temp_list = []
    for row in range(3,100):
        check_empty = ws[get_column_letter(1) + str(row)].value
        if check_empty is None:
            break
        for col in range(1,7):
            temp_list.append(ws[get_column_letter(col) + str(row)].value)
    print(temp_list)
    return temp_list

            
def calendar_gui():
    """_summary_
    """
    root = Tk()

    display_date = get_date()
    my_label = Label(root, text=('Date', display_date))
    my_label.place(relx=0.0, rely=0.0, anchor='nw')

    def displayed_events():
        display_todays_event = check_todays_events()
        todays_events_frame = LabelFrame(root, text="Todays Events")
        todays_events_frame.place(width=300, height=100, x=200, y=200)
        scrollbar_y = Scrollbar(todays_events_frame)
        scrollbar_y.pack(side=RIGHT, fill=Y)
        todays_events_text = Text(todays_events_frame, width = 15, height = 15, wrap = NONE,
                    yscrollcommand = scrollbar_y.set)
        for event in display_todays_event:
            string = display_todays_event[display_todays_event.index(event)] + "\n"
            todays_events_text.insert(END, string)
        todays_events_text.config(state=DISABLED)
        todays_events_text.pack(side=TOP, fill=X)
        scrollbar_y.config(command=todays_events_text.yview)

        display_upcoming_event = check_upcoming_events()
        upcoming_events_frame = LabelFrame(root, text="Upcoming Events")
        upcoming_events_frame.place(width=300, height=100, x=200, y=300)
        scrollbar_y = Scrollbar(upcoming_events_frame)
        scrollbar_y.pack(side=RIGHT, fill=Y)
        upcoming_events_text = Text(upcoming_events_frame, width = 15, height = 15, wrap = NONE,
                    yscrollcommand = scrollbar_y.set)
        for event in display_upcoming_event[0]:
            string = display_upcoming_event[0][display_upcoming_event[0].index(event)] + " in " + display_upcoming_event[1][display_upcoming_event[0].index(event)] + " Day(s)\n"
            upcoming_events_text.insert(END, string)
        upcoming_events_text.config(state=DISABLED)
        upcoming_events_text.pack(side=TOP, fill=X)
        scrollbar_y.config(command=upcoming_events_text.yview)
        return upcoming_events_text
    
    displayed_events()

    appointment_label = Label(root, text=('Create an Appointment/Event'))
    appointment_label.place(x=38, y=35)

    delete_label = Label(root, text='Delete an Appointment/Event')
    delete_label.place(x=329, y=35)

    invalid_label = Label(root)
    my_label_added_event = Label(root)

    invalid_delete_label = Label(root)
    my_label_deleted_event = Label(root)

    mini_calendar_window = Toplevel()
    mini_calendar_window.withdraw()

    def do_an_entry():
        valid_day = True
        day_value = entry_day.get()
        valid_month = True
        month_value = entry_month.get()
        valid_year = True
        year_value = entry_year.get()
        description_value = entry_description.get()
        valid_reminder = True
        reminder_value = entry_reminder.get()
        
        if day_value == "":
            print("Invalid input")
            valid_day = False
        for character in day_value:
            try:
                converted_character = int(character)
                if converted_character not in range(0, 10) or int(day_value) > 31:
                    valid_day = False
                if len(day_value) > 2:
                    valid_day = False
            except ValueError:
                valid_day = False
                pass
        
        if month_value == "":
            valid_month = False
        for character in month_value:
            try:
                converted_character = int(character)
                if converted_character not in range(0, 10) or int(month_value) > 12:
                    valid_month = False
                if len(month_value) > 2 or len(month_value) == 0 or month_value is None:
                    valid_month = False
            except ValueError:
                valid_month = False
                pass
        
        if year_value == "":
            valid_year = False
        for character in year_value:
            try:
                converted_character = int(character)
                if converted_character not in range(0, 10):
                    valid_year = False
                if len(year_value) > 4 or len(year_value) < 2 or len(year_value) == 3 or len(year_value) == 0 or not year_value:
                    valid_year = False
            except ValueError:
                valid_year = False
                pass
        
        if reminder_value == "":
            reminder_value = "0"
        for character in reminder_value:
            try:
                converted_character = int(character)
                if converted_character not in range(0, 10):
                    valid_reminder = False
            except ValueError:
                valid_reminder = False
                pass

        if valid_day is True and valid_month is True and valid_year is True and valid_reminder is True:
            success_label = f'Your event "{description_value}" for the {day_value}.{month_value}.{year_value} was successfully added\nYou will get notified {reminder_value} days before'
            messagebox.showinfo("Event Added", success_label)
            add_event(day_value, month_value, year_value, description_value, reminder_value)
        else:
            messagebox.showerror("Event could not be Added", "Invalid Input: Check your Entries again")

    def delete_an_entry():
        valid_id = True
        id_value = entry_id.get()

        for character in id_value:
            try:
                converted_character = int(character)
                if converted_character not in range(0, 10):
                    valid_id = False
            except ValueError:
                valid_day = False
                pass

        if valid_id is True:
            wb = load_workbook('my_Events.xlsx')
            ws = wb["all_Events"]
            temp_list = []
            for row in range(3,100):
                check_id = ws[get_column_letter(1) + str(row)].value
                if check_id is None:
                    print(f"Event with the id {id_value} could not be found")
                    break
                if check_id == id_value:
                    temp_list.append(ws[get_column_letter(2) + str(row)].value)
                    temp_list.append(ws[get_column_letter(3) + str(row)].value)
                    temp_list.append(ws[get_column_letter(4) + str(row)].value)
                    temp_list.append(ws[get_column_letter(5) + str(row)].value)
                    break
            if temp_list == []:
                messagebox.showerror("Event could not be Deleted", f"Event with the id: {id_value}\ncould not be found")
            else:
                messagebox.showinfo("Event could not be Deleted", f"Your event {temp_list[3]}\nfor the {temp_list[0]}.{temp_list[1]}.{temp_list[2]}\nwas successfully deleted")
                del_event(id_value)
            
    def event_text(e):
        entry_day.delete(0, "end")
        entry_month.delete(0, "end")
        entry_year.delete(0, "end")
        entry_description.delete(0, "end")
        entry_reminder.delete(0, "end")
    
    def delete_text(e):
        entry_id.delete(0, "end")

    # Form to create an Event
    label_day = Label(root, text="Day")
    label_day.place(x=5, y=60)
    entry_day = Entry(root)
    entry_day.insert(0,"11")                                # Preview Text
    entry_day.place(x=75, y=60)
    label_month = Label(root, text="Month")
    label_month.place(x=5, y=82)
    entry_month = Entry(root)
    entry_month.insert(0, "09")                             # Preview Text
    entry_month.place(x=75, y=82)
    label_year = Label(root, text="Year")
    label_year.place(x=5, y=104)
    entry_year = Entry(root)
    entry_year.insert(0, "2064")                            # Preview Text
    entry_year.place(x=75, y=104)
    label_description = Label(root, text="Description")
    label_description.place(x=5, y=126)
    entry_description = Entry(root)                                             # ToDo change to TextBox
    entry_description.insert(0, "Dennis 63th Bday")         # Preview Text
    entry_description.place(height=40 ,x=75, y=126)
    label_reminder = Label(root, text="Reminder")
    label_reminder.place(x=5, y=168)
    entry_reminder = Entry(root)
    entry_reminder.insert(0, "5")                           # Preview Text
    entry_reminder.place(width=40, x=75, y= 168)

    submit_event = Button(root, text="Add Event", padx=8, pady=3, command=do_an_entry)
    submit_event.place(x=75, y=193)

    entry_day.bind("<FocusIn>", event_text)

    def input_date_menu():
        if 'normal' == mini_calendar_window.state():
            mini_calendar_window.withdraw()
        else:
            mini_calendar_window.deiconify()
            mini_calendar_temp_label = Calendar(mini_calendar_window, selectmode="day", year=2022, month=10, day=20)
            mini_calendar_temp_label.pack()
            mini_calendar_window.geometry("250x212")
            mini_calendar_window.overrideredirect(True)

            def grab_date():
                selected_date = mini_calendar_temp_label.get_date()
                selected_day = selected_date[0] + selected_date[1]
                selected_month = selected_date[3] + selected_date[4]
                selected_year = selected_date[6] + selected_date[7]

                entry_day.delete(0,'end')
                entry_month.delete(0,'end')
                entry_year.delete(0,'end')
                entry_day.insert(0,selected_day)
                entry_month.insert(0,selected_month)
                entry_year.insert(0,selected_year)

            Button(mini_calendar_window, text="get Date", command=grab_date).place(width=250, x=0, y=186)

            myLabel_test = Label(mini_calendar_window, text="")
            myLabel_test.pack(pady=20)



    input_date_btn = Button(root)
    photo_calender = PhotoImage(file="calendar_button.png")
    input_date_btn.config(image= photo_calender, width= "17", height= "16", activeforeground= "black", bg= "black", bd=0, command=input_date_menu)
    input_date_btn.place(x=205, y=105)

    def refresh_events():
        """_summary_
        """
        displayed_events()

    refresh_button = Button(root)
    photo_refresh = PhotoImage(file="Refresh_icon.png")
    refresh_button.config(image= photo_refresh, width="14", height="15", activeforeground= "black", bg= "black", bd=0, command=refresh_events)
    refresh_button.place(x=480, y=0)


    # Form to delete an Event
    label_id = Label(root, text="Event ID")
    label_id.place(x=320, y=60)
    entry_id = Entry(root)
    entry_id.insert(0,"3067211771744094029")                                # Preview Text
    entry_id.place(x=370, y=60)

    submit_delete = Button(root, text="Delete Event", padx=8, pady=3, command=delete_an_entry)
    submit_delete.place(x=400, y=85)

    entry_id.bind("<FocusIn>", delete_text)

    def show_events():
        event_window = Toplevel()
        event_window.title('Calendar/Events')
        event_window.iconbitmap('Apple_Calendar_Icon.ico')
        
        events = show_all_events()
        num_of_events = len(events)//6
        all_events_frame = Frame(event_window, width=800, height=500)
        all_events_frame.pack(fill=BOTH, expand=1)
        all_events_canvas = Canvas(all_events_frame, width=800, height=500)
        all_events_canvas.pack(side=LEFT, fill=BOTH, expand=1)
        scrollbar_y = Scrollbar(all_events_frame, orient=VERTICAL, command=all_events_canvas.yview)
        scrollbar_y.pack(side=RIGHT, fill=Y)
        all_events_canvas.configure(yscrollcommand=scrollbar_y.set)
        all_events_canvas.bind('<Configure>', lambda e: all_events_canvas.configure(scrollregion=all_events_canvas.bbox("all")))
        all_events = Text(all_events_canvas, width=800, height=500)
        for event in range(1, num_of_events+1):
            temp = str("ID: " + events[0]) + "  Day: " + str(events[1]) + "  Month: " + str(events[2]) + "  Year: " + str(events[3]) + "  Description: " + str(events[4]) + "  Reminder: " + str(events[5] + "\n")
            del events[:6]
            all_events.insert(END, temp)
        all_events.config(state=DISABLED)
        all_events.pack(side=TOP, fill=X)
        all_events_canvas.create_window((0,0), window=all_events, anchor="nw")

    def past_events():
        event_window = Toplevel()
        event_window.title('Calendar/Events')
        event_window.iconbitmap('Apple_Calendar_Icon.ico')
        
        events = show_past_events()
        num_of_events = len(events)//6
        past_events_frame = Frame(event_window, width=800, height=500)
        past_events_frame.pack(fill=BOTH, expand=1)
        past_events_canvas = Canvas(past_events_frame, width=800, height=500)
        past_events_canvas.pack(side=LEFT, fill=BOTH, expand=1)
        scrollbar_y = Scrollbar(past_events_frame, orient=VERTICAL, command=past_events_canvas.yview)
        scrollbar_y.pack(side=RIGHT, fill=Y)
        past_events_canvas.configure(yscrollcommand=scrollbar_y.set)
        past_events_canvas.bind('<Configure>', lambda e: past_events_canvas.configure(scrollregion=past_events_canvas.bbox("all")))
        past_events = Text(past_events_canvas, width=800, height=500)
        for event in range(1, num_of_events+1):
            temp = str("ID: " + events[0]) + "  Day: " + str(events[1]) + "  Month: " + str(events[2]) + "  Year: " + str(events[3]) + "  Description: " + str(events[4]) + "  Reminder: " + str(events[5] + "\n")
            del events[:6]
            past_events.insert(END, temp)
        past_events.config(state=DISABLED)
        past_events.pack(side=TOP, fill=X)
        past_events_canvas.create_window((0,0), window=past_events, anchor="nw")

    def upcoming_events():
        event_window = Toplevel()
        event_window.title('Calendar/Events')
        event_window.iconbitmap('Apple_Calendar_Icon.ico')
        
        events = show_upcoming_events()
        num_of_events = len(events)//6
        upcoming_events_frame = Frame(event_window, width=800, height=500)
        upcoming_events_frame.pack(fill=BOTH, expand=1)
        upcoming_events_canvas = Canvas(upcoming_events_frame, width=800, height=500)
        upcoming_events_canvas.pack(side=LEFT, fill=BOTH, expand=1)
        scrollbar_y = Scrollbar(upcoming_events_frame, orient=VERTICAL, command=upcoming_events_canvas.yview)
        scrollbar_y.pack(side=RIGHT, fill=Y)
        upcoming_events_canvas.configure(yscrollcommand=scrollbar_y.set)
        upcoming_events_canvas.bind('<Configure>', lambda e: upcoming_events_canvas.configure(scrollregion=upcoming_events_canvas.bbox("all")))
        upcoming_events = Text(upcoming_events_canvas, width=800, height=500)
        for event in range(1, num_of_events+1):
            temp = str("ID: " + events[0]) + "  Day: " + str(events[1]) + "  Month: " + str(events[2]) + "  Year: " + str(events[3]) + "  Description: " + str(events[4]) + "  Reminder: " + str(events[5] + "\n")
            del events[:6]
            upcoming_events.insert(END, temp)
        upcoming_events.config(state=DISABLED)
        upcoming_events.pack(side=TOP, fill=X)
        upcoming_events_canvas.create_window((0,0), window=upcoming_events, anchor="nw")

    show_all_events_button = Button(root, text="Show all events", padx=8, pady=3, command=show_events)
    show_all_events_button.place(width=155,x=20, y=260)

    show_upcoming_events_button = Button(root, text="Show upcoming events", padx=8, pady=3, command=upcoming_events)
    show_upcoming_events_button.place(width=155, x=20, y=292)

    show_past_events_button = Button(root, text="Show past events", padx=8, pady=3, command=past_events)
    show_past_events_button.place(width=155, x=20, y=324)

    # Create Windows Notification
    upcoming_events_text = displayed_events()
    input_from_textbox = upcoming_events_text.get("1.0", END)
    if len(input_from_textbox) == 1:
        pass
    else:
        toast = Notification(app_id="myNotification",
                        title="You Have Upcoming Events!:",
                        msg=input_from_textbox,
                        duration="long",
                        icon=path.abspath(r"Apple_Calendar_Icon.ico"))
        toast.set_audio(audio.Default, loop=False)
        toast.show()

    root.title('Calendar')
    root.iconbitmap('Apple_Calendar_Icon.ico')
    root.geometry("500x400")
    root.mainloop()

write_past_events()
write_upcoming_events()
calendar_gui()

# if (__name__ == "__main__"):
#      main()
